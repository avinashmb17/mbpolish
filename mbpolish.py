import pandas as pd
import streamlit as st
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows



#import pandas as pd
#import streamlit as st
#from io import BytesIO
#from openpyxl import Workbook
#from openpyxl.styles import (
#    Font,
#    Border,
#    Side,
#    PatternFill,
#    Alignment
#)
#from openpyxl.utils.dataframe import dataframe_to_rows

#from openpyxl import Workbook
#from openpyxl.styles import Font, Border, Side
#from openpyxl.utils.dataframe import dataframe_to_rows

# ---------------- PAGE SETUP ---------------- #

def setup_page():
    st.set_page_config(
        page_title="Polish Details",
        page_icon=":bar_chart:",
        layout="wide"
    )


# ---------------- FILE UPLOAD ---------------- #

def file_upload():
    return st.sidebar.file_uploader(
        "Choose Data File",
        type=['xls', 'xlsx']
    )


def mst_upload():
    return st.sidebar.file_uploader(
        "Choose Master Data File",
        type=['xls', 'xlsx']
    )


# ---------------- LOT FILTER ---------------- #

def lot(df_data):

    lot_list = (
        df_data['LOT']
        .dropna()
        .astype(str)
        .unique()
        .tolist()
    )

    selected_lot = st.sidebar.multiselect(
        "Select Material LOT",
        options=lot_list
    )
    
    

    return selected_lot


# ---------------- CACHE: MASTER FILE ---------------- #

@st.cache_data
def load_master(mst):

    df_dtc_quality = pd.read_excel(
        mst,
        sheet_name="master",
        usecols='A:C',
        skiprows=2
    ).dropna(how='all')

    df_shape_mst = pd.read_excel(
        mst,
        sheet_name="master",
        usecols='H:J',
        skiprows=2
    ).dropna(how='all')

    df_shade_mst = pd.read_excel(
        mst,
        sheet_name="master",
        usecols='L:Q',
        skiprows=2
    ).dropna(how='all')

    df_size_mst = pd.read_excel(
        mst,
        sheet_name="master",
        usecols='S:X',
        skiprows=2
    ).dropna(how='all')

    df_mbgd_P08_mst = pd.read_excel(
        mst,
        sheet_name="master",
        usecols='AD:AF',
        skiprows=2
    ).dropna(how='all')

    df_mbgd_mst = pd.read_excel(
        mst,
        sheet_name="master",
        usecols='Z:AF',
        skiprows=2
    ).dropna(how='all')

    df_dtc_size = pd.read_excel(
        mst,
        sheet_name="master",
        usecols='AH:AI',
        skiprows=2
    ).dropna(how='all')

    return (
        df_dtc_quality,
        df_shape_mst,
        df_shade_mst,
        df_size_mst,
        df_mbgd_mst,
        df_mbgd_P08_mst,
        df_dtc_size
    )


# ---------------- CACHE: DATA FILE ---------------- #

@st.cache_data
def load_data(file):

    df_data = pd.read_excel(
        file,
        sheet_name="FORMAT",
        header=8
    ).dropna(how='all')

    df_data.columns = df_data.columns.str.strip()

    df_data = df_data.loc[
        :,
        ~df_data.columns.astype(str).str.contains('^Unnamed')
    ]

    df_data.reset_index(drop=True, inplace=True)

    return df_data


# ---------------- MERGE FUNCTIONS ---------------- #

def dtc_shape(df_data, df_shape_mst):
    df_shape_mst = df_shape_mst.drop_duplicates()

    df_data = df_data.merge(
        df_shape_mst[['SHAPE', 'NEW SHAPE', 'NEW SHAPE-ID']],
        on='SHAPE',
        how='left'
    )

    df_data.rename(columns={
        'NEW SHAPE': 'DTC SHAPE',
        'NEW SHAPE-ID': 'DTC SHAPE-ID'
    }, inplace=True)

    return df_data


def dtc_quality(df_data, df_dtc_quality):
    df_dtc_quality = df_dtc_quality.drop_duplicates()
    return df_data.merge(
        df_dtc_quality[['DTC-QLY.', 'DTC QLTY', 'DTC QLTY-ID']],
        on='DTC-QLY.',
        how='left'
    )


def shade(df_data, df_shade_mst):
    df_shade_mst =df_shade_mst.drop_duplicates()
    df_data = df_data.merge(
        df_shade_mst[['COL', 'SHADE', 'DEPT.', 'SHADE-II', 'SHADE-II-ID']],
        on=['COL', 'SHADE', 'DEPT.'],
        how='left'
    )

    df_data.rename(columns={
        'SHADE-II': 'DTC SHADE',
        'SHADE-II-ID': 'DTC SHADE-ID'
    }, inplace=True)

    return df_data


def mb_size(df_data, df_size_mst):
    df_size_mst = df_size_mst.drop_duplicates()
    df_data.columns = df_data.columns.str.strip()
    df_size_mst.columns = df_size_mst.columns.str.strip()

    df_data = df_data.merge(
        df_size_mst[
            ['To Material Group', 'CLARITY', 'SIZE',
             'MB_SIZE_GR', 'MB_SIZE_GR-ID']
        ],
        left_on=['POLISH LOT', 'IGI QLTY', 'SIZE-H'],
        right_on=['To Material Group', 'CLARITY', 'SIZE'],
        how='left'
    )

    df_data.drop(columns=[
        'To Material Group'
    ], errors='ignore', inplace=True)

    df_data.loc[
        df_data['POLISH LOT'].isin(['P-08', 'P-09']),
        'MB_SIZE_GR'
    ] = 'Cert & Noncert'

    df_data.loc[
        df_data['POLISH LOT'] == 'P-07',
        'MB_SIZE_GR'
    ] = 'Fancy'

    return df_data

def dtc_size(df_data, df_dtc_size_mst):
    df_dtc_size_mst = df_dtc_size_mst.drop_duplicates()
    df_data.columns = df_data.columns.str.strip()
    df_dtc_size_mst.columns = df_dtc_size_mst.columns.str.strip()

    df_data = df_data.merge(
        df_dtc_size_mst[
            ['DTC-SIZE',
             'DTC-SIZE-ID']
        ],
        left_on=['DTC SIZE'],
        right_on=['DTC-SIZE'],
        how='left'
    )
    return df_data



def mb_gdstatus(df_data, df_mbgd_mst, df_mbgd_P08_mst):
    df_mbgd_mst = df_mbgd_mst.drop_duplicates()
    df_mbgd_P08_mst = df_mbgd_P08_mst.drop_duplicates()
    df_data.columns = df_data.columns.str.strip()
    df_mbgd_mst.columns = df_mbgd_mst.columns.str.strip()
    df_mbgd_P08_mst.columns = df_mbgd_P08_mst.columns.str.strip()

    df_mbgd_mst = df_mbgd_mst.drop_duplicates()
    df_mbgd_P08_mst = df_mbgd_P08_mst.drop_duplicates()

    df_data['MATCH_KEY'] = (
        df_data['POLISH LOT'].astype(str).str.strip()
        + df_data['DTC SHADE'].astype(str).str.strip()
        + df_data['QUALITY'].astype(str).str.strip()
        + df_data['SIZE-H'].astype(str).str.strip()
    )

    df_mbgd_mst['MATCH_KEY'] = (
        df_mbgd_mst['POLISH LOT'].astype(str).str.strip()
        + df_mbgd_mst['DTC-SHADE'].astype(str).str.strip()
        + df_mbgd_mst['MB_CLARITY'].astype(str).str.strip()
        + df_mbgd_mst['SIZE.1'].astype(str).str.strip()
    )

    df_data = df_data.merge(
        df_mbgd_mst[['MATCH_KEY', 'MB-GD-STATUS', 'MB-GD-STATUS-ID']],
        on='MATCH_KEY',
        how='left'
    )

    df_data['MATCH_KEY_P08'] = (
        df_data['POLISH LOT'].astype(str).str.strip()
        + df_data['Remarks'].astype(str).str.strip()
    )

    df_data = df_data.merge(
        df_mbgd_P08_mst[['CODE1', 'MB-GD-STATUS', 'MB-GD-STATUS-ID']],
        left_on='MATCH_KEY_P08',
        right_on='CODE1',
        how='left',
        suffixes=('', '_P08')
    )

    p08_mask = df_data['POLISH LOT'] == 'P-08'

    df_data.loc[p08_mask, 'MB-GD-STATUS'] = df_data.loc[p08_mask, 'MB-GD-STATUS_P08']
    df_data.loc[p08_mask, 'MB-GD-STATUS-ID'] = df_data.loc[p08_mask, 'MB-GD-STATUS-ID_P08']

    df_data.fillna({'MB-GD-STATUS': 'MIX', 'MB-GD-STATUS-ID': '01'}, inplace=True)

    df_data.drop(columns=[
        'MATCH_KEY', 'MATCH_KEY_P08', 'CODE1',
        'MB-GD-STATUS_P08', 'MB-GD-STATUS-ID_P08'
    ], errors='ignore', inplace=True)

    return df_data

#Export to Excel
def export_to_excel(
    pivot1,
    pivot2,
    pivot3,
    pivot4,
    pivot5,
    selected_lot
):

    output = BytesIO()

    wb = Workbook()
    ws = wb.active
    ws.title = "Polish Report"

    # ====================================
    # STYLES
    # ====================================

    bold_font = Font(
        bold=True
    )

    header_fill = PatternFill(
        start_color="D9D9D9",
        end_color="D9D9D9",
        fill_type="solid"
    )

    total_fill = PatternFill(
        start_color="FFF2CC",
        end_color="FFF2CC",
        fill_type="solid"
    )

    thin = Side(
        border_style="thin",
        color="000000"
    )

    border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin
    )

    # Text Alignment
    text_align = Alignment(
        horizontal="left",
        vertical="center"
    )

    # Number Alignment
    number_align = Alignment(
        horizontal="right",
        vertical="center"
    )

    # ====================================
    # MAIN HEADER
    # ====================================

    ws["A1"] = f"Material Number : {', '.join(selected_lot)}"
    ws["A1"].font = Font(
        bold=True,
        size=14
    )

    ws["A2"] = "Article Name : "
    ws["A2"].font = Font(
        bold=True,
        size=12
    )

    # ====================================
    # WRITE PIVOT FUNCTION
    # ====================================

    def write_pivot(df, start_row, start_col, title):

        # ----------------------------
        # Pivot Title
        # ----------------------------

        title_cell = ws.cell(
            row=start_row,
            column=start_col
        )

        title_cell.value = title

        title_cell.font = Font(
            bold=True,
            size=12
        )

        # ----------------------------
        # Write DataFrame
        # ----------------------------

        for r_idx, row in enumerate(
            dataframe_to_rows(
                df,
                index=False,
                header=True
            ),
            start_row + 1
        ):

            row_values = [
                str(v).upper() if v is not None else ""
                for v in row
            ]

            is_total_row = (
                "TOTAL" in row_values
            )

            for c_idx, value in enumerate(
                row,
                start_col
            ):

                cell = ws.cell(
                    row=r_idx,
                    column=c_idx
                )

                cell.value = value

                # ======================
                # BORDER
                # ======================

                cell.border = border

                # ======================
                # HEADER FORMAT
                # ======================

                if r_idx == start_row + 1:

                    cell.font = Font(
                        bold=True
                    )

                    cell.fill = header_fill

                # ======================
                # TOTAL ROW FORMAT
                # ======================

                if is_total_row:

                    cell.font = Font(
                        bold=True
                    )

                    cell.fill = total_fill

                # ======================
                # ALIGNMENT
                # ======================

                if isinstance(value, (int, float)):

                    cell.alignment = number_align

                    cell.number_format = '#,##0.00'

                else:

                    cell.alignment = text_align

    # ====================================
    # WRITE PIVOTS
    # ====================================

    write_pivot(
        pivot1,
        6,
        1,
        "Pivot1"
    )

    write_pivot(
        pivot2,
        6,
        10,
        "Pivot2"
    )

    write_pivot(
        pivot3,
        30,
        1,
        "Pivot3"
    )

    write_pivot(
        pivot4,
        30,
        10,
        "Pivot4"
    )

    write_pivot(
        pivot5,
        55,
        1,
        "Pivot5"
    )

    # ====================================
    # AUTO COLUMN WIDTH
    # ====================================

    for column_cells in ws.columns:

        max_length = 0

        column_letter = column_cells[0].column_letter

        for cell in column_cells:

            try:

                max_length = max(
                    max_length,
                    len(str(cell.value))
                )

            except:
                pass

        ws.column_dimensions[
            column_letter
        ].width = max_length + 4

    # ====================================
    # REMOVE DEFAULT GRIDLINES
    # ====================================

    ws.sheet_view.showGridLines = False

    # ====================================
    # SAVE FILE
    # ====================================

    wb.save(output)

    output.seek(0)

    return output
# ---------------- Pivot Report Tables ---------------- #

def create_pivots(df_data,selected_lot) -> None:

    # ==========================================
    # FILTER ONLY ROUND SHAPE
    # ==========================================

    df_round = df_data[
        df_data['DTC SHAPE'].astype(str).str.upper() == 'ROUND'
    ].copy()

    # ==========================================
    # CALCULATIONS
    # ==========================================
    # Showing Data format final file showing hide to screen 
    #st.dataframe(df_round)
    #st.write(df_round.columns.tolist())
    
    df_round['VALUE/CTS'] = (
        df_round['VALUE'] / df_round['CTS']
    )

    
    #total_cts = df_round['CTS'].sum()
    #total_value = df_round['VALUE'].sum()
    #avg_value_cts = total_value / total_cts
    total_cts = round(df_round['CTS'].sum(), 2)
    total_value = round(df_round['VALUE'].sum(), 2)
    avg_value_cts = round(total_value / total_cts, 2)
    
   
   
   # =====================================================
    # SORTING IDS
    # =====================================================

    # convert ids numeric if needed
    df_round["MB_GD_STATUS-ID"] = pd.to_numeric(
        df_round["MB-GD-STATUS-ID"],
        errors="coerce"
    )
        

    df_round["DTC SHAPE-ID"] = pd.to_numeric(
        df_round["DTC SHAPE-ID"],
        errors="coerce"
    )
   
    df_round["DTC SHADE-ID"] =  pd.to_numeric(
       df_round["DTC SHADE-ID"],
       errors = "coerce"
   )
   
    # STREAMLIT COLUMNS
    # =====================================================

    #col1, col2, col3 = st.columns(3)
    col1, col2= st.columns(2)
    col3, col4= st.columns(2)
    col5, col6= st.columns(2)
    #col5 = st.columns(1)[0]
    # =====================================================
    # PIVOT TABLE 1
    # MB_GD_STATUS + DTC QLTY
    # ==========================
# STEP 0: CLEAN DATA
# ==========================

    df_round["MB-GD-STATUS"] = (
    df_round["MB-GD-STATUS"]
    .astype(str)
    .str.replace("\u00A0", " ")   # remove hidden spaces
    .str.strip()
    .str.upper()
)

    df_round["DTC QLTY"] = (
    df_round["DTC QLTY"]
    .astype(str)
    .str.strip()
)

    # REMOVE OLD TOTAL ROWS IF EXIST
    df_clean = df_round[
        ~df_round["DTC QLTY"].astype(str).str.upper().eq("TOTAL")
    ].copy()

    # ==========================
    # STEP 1: ID MAP FOR SORTING
    # ==========================

    id_map = df_clean[[
        "MB-GD-STATUS",
        "DTC QLTY",
        "MB_GD_STATUS-ID",
        "DTC QLTY-ID"
    ]].drop_duplicates()

    # ==========================
    # STEP 2: BASE AGGREGATION
    # ==========================

    base = df_clean.groupby(
            ["MB-GD-STATUS", "DTC QLTY"],
            as_index=False
        )[["CTS", "VALUE"]].sum()

        # merge IDs
    base = base.merge(
            id_map,
            on=["MB-GD-STATUS", "DTC QLTY"],
            how="left"
        )

        # sort properly
    base = base.sort_values(
            by=["MB_GD_STATUS-ID", "DTC QLTY-ID"],
            na_position="last"
        )

        # ==========================
        # STEP 3: BUILD STRUCTURE
        # ==========================

    final_rows = []

    group_order = base.sort_values("MB_GD_STATUS-ID")["MB-GD-STATUS"].unique()

    for grp in group_order:

            grp_df = base[base["MB-GD-STATUS"] == grp].copy()

            # detail rows
            final_rows.append(grp_df[[
                "MB-GD-STATUS",
                "DTC QLTY",
                "CTS",
                "VALUE"
            ]])

            # subtotal row
            subtotal = pd.DataFrame({
                "MB-GD-STATUS": [grp],
                "DTC QLTY": ["Total"],
                "CTS": [grp_df["CTS"].sum()],
                "VALUE": [grp_df["VALUE"].sum()]
            })

            final_rows.append(subtotal)

        # ==========================
        # STEP 4: COMBINE
        # ==========================

    pivot1 = pd.concat(final_rows, ignore_index=True)

        # ==========================
        # STEP 5: GRAND TOTAL (IMPORTANT FIX)
        # ==========================

    real_total_cts = df_clean["CTS"].sum()
    real_total_value = df_clean["VALUE"].sum()

    grand_total = pd.DataFrame({
            "MB-GD-STATUS": ["TOTAL"],
            "DTC QLTY": [""],
            "CTS": [real_total_cts],
            "VALUE": [real_total_value]
        })

    pivot1 = pd.concat([pivot1, grand_total], ignore_index=True)

    # ==========================
    # STEP 6: CALCULATIONS
    # ==========================

    pivot1["CTS %"] = (pivot1["CTS"] / real_total_cts * 100).round(2)
    pivot1["VALUE %"] = (pivot1["VALUE"] / real_total_value * 100).round(2)

    pivot1["AVG VALUE/CTS"] = 0

    mask = pivot1["CTS"] != 0

    pivot1.loc[mask, "AVG VALUE/CTS"] = (
        pivot1.loc[mask, "VALUE"] /
        pivot1.loc[mask, "CTS"]
        ).round(2)

    # ==========================
    # STEP 7: FINAL COLUMN ORDER
    # ==========================

    pivot1 = pivot1[[
            "MB-GD-STATUS",
            "DTC QLTY",
            "CTS",
            "CTS %",
            "VALUE",
            "VALUE %",
            "AVG VALUE/CTS"
        ]]
            

    # =====================================================
    # PIVOT TABLE 2
    # MB_SIZE_GR
    # =====================================================

    pivot2 = pd.pivot_table(
        df_round,
        index=[
            "MB_SIZE_GR",
        ],
        values=[
            "CTS",
            "VALUE"
        ],
        aggfunc={
            "CTS": "sum",
            "VALUE": "sum"
        },
        margins=True,
        margins_name="TOTAL"
    ).reset_index()

    # CTS %
    pivot2["CTS %"] = (
        pivot2["CTS"] / total_cts * 100
    ).round(2)

    # VALUE %
    pivot2["VALUE %"] = (
        pivot2["VALUE"] / total_value * 100
    ).round(2)

    # AVG VALUE/CTS
    pivot2["AVG VALUE/CTS"] = (
        pivot2["VALUE"] / pivot2["CTS"]
    ).round(2)
    
     # Round numeric columns to 2 decimals
    pivot2[["CTS", "CTS %", "VALUE", "VALUE %", "AVG VALUE/CTS"]] = (
    pivot2[["CTS", "CTS %", "VALUE", "VALUE %", "AVG VALUE/CTS"]].round(2)
    )

    # SORTING
    sort_df2 = (
        df_round[[
            "MB_SIZE_GR",
            "MB_SIZE_GR-ID"
        ]]
        .drop_duplicates()
    )

    pivot2 = pivot2.merge(
        sort_df2,
        on=[
            "MB_SIZE_GR",
                    ],
        how="left"
    )

    pivot2 = pivot2.sort_values(
        by=[
            "MB_SIZE_GR-ID"
        ],
        na_position="last"
    )

    # COLUMN ORDER
    pivot2 = pivot2[[
        "MB_SIZE_GR",
        "CTS",
        "CTS %",
        "VALUE",
        "VALUE %",
        "AVG VALUE/CTS"
    ]]


# =====================================================
    # PIVOT TABLE 3
    # DTC Shade
    # =====================================================

    pivot3 = pd.pivot_table(
        df_round,
        index=[
            "DTC SHADE",
            "DTC SHADE-ID",    
        ],
        values=[
            "CTS",
            "VALUE"
        ],
        aggfunc={
            "CTS": "sum",
            "VALUE": "sum"
        },
        margins=True,
        margins_name="TOTAL"
    ).reset_index()

    # CTS %
    pivot3["CTS %"] = (
        pivot3["CTS"] / total_cts * 100
    ).round(2)

    # VALUE %
    pivot3["VALUE %"] = (
        pivot3["VALUE"] / total_value * 100
    ).round(2)

    # AVG VALUE/CTS
    pivot3["AVG VALUE/CTS"] = (
        pivot3["VALUE"] / pivot3["CTS"]
    ).round(2)
    
     # Round numeric columns to 2 decimals
    pivot3[["CTS", "CTS %", "VALUE", "VALUE %", "AVG VALUE/CTS"]] = (
    pivot3[["CTS", "CTS %", "VALUE", "VALUE %", "AVG VALUE/CTS"]].round(2)
    )

    pivot3["DTC SHADE-ID"] = pd.to_numeric(
    pivot3["DTC SHADE-ID"],
    errors="coerce"
    )

    # SORTING
    sort_df3 = (
        df_round[[
            "DTC SHADE",
            "DTC SHADE-ID",
        ]]
        .drop_duplicates()
    )
    
#    pivot3 = pivot3.merge(
#        sort_df3,
#        on=[
#            "DTC SHADE-ID",
#                    ],
#        how="left"
#    )

    pivot3 = pivot3.sort_values(
        by=[
            "DTC SHADE-ID"
        ],
        na_position="last"
    )

    # COLUMN ORDER
    pivot3 = pivot3[[
        "DTC SHADE",
        "CTS",
        "CTS %",
        "VALUE",
        "VALUE %",
        "AVG VALUE/CTS"
    ]]

# =====================================================
    # PIVOT TABLE 4
    # DTC Size
    # =====================================================

    pivot4 = pd.pivot_table(
        df_round,
        index=[
            "DTC SIZE",
            "DTC-SIZE-ID",    
        ],
        values=[
            "CTS",
            "VALUE"
        ],
        aggfunc={
            "CTS": "sum",
            "VALUE": "sum"
        },
        margins=True,
        margins_name="TOTAL"
    ).reset_index()

    # CTS %
    pivot4["CTS %"] = (
        pivot4["CTS"] / total_cts * 100
    ).round(2)

    # VALUE %
    pivot4["VALUE %"] = (
        pivot4["VALUE"] / total_value * 100
    ).round(2)

    # AVG VALUE/CTS
    pivot4["AVG VALUE/CTS"] = 0
    mask = pivot4["CTS"] != 0

    pivot4.loc[mask, "AVG VALUE/CTS"] = (
    pivot4.loc[mask, "VALUE"] /
    pivot4.loc[mask, "CTS"]
        ).round(2)
    
    # AVG VALUE/CTS
    #pivot4["AVG VALUE/CTS"] = (
    #    pivot4["VALUE"] / pivot3["CTS"]
    #).round(2)
    
     # Round numeric columns to 2 decimals
    pivot4[["CTS", "CTS %", "VALUE", "VALUE %", "AVG VALUE/CTS"]] = (
    pivot4[["CTS", "CTS %", "VALUE", "VALUE %", "AVG VALUE/CTS"]].round(2)
    )

    pivot4["DTC-SIZE-ID"] = pd.to_numeric(
    pivot4["DTC-SIZE-ID"],
    errors="coerce"
    )

    # SORTING
    sort_df4 = (
        df_round[[
            "DTC SIZE",
            "DTC-SIZE-ID",
        ]]
        .drop_duplicates()
    )

    pivot4 = pivot4.sort_values(
        by=[
            "DTC-SIZE-ID"
        ],
        na_position="last"
    )

    # COLUMN ORDER
    pivot4 = pivot4[[
        "DTC SIZE",
        "CTS",
        "CTS %",
        "VALUE",
        "VALUE %",
        "AVG VALUE/CTS"
    ]]

# =====================================================
    # PIVOT TABLE 5
    # DTC Shape
    # =====================================================

    pivot5 = pd.pivot_table(
        df_data,
        index=[
            "DTC SHAPE",
            "DTC SHAPE-ID",    
        ],
        values=[
            "CTS",
            "VALUE"
        ],
        aggfunc={
            "CTS": "sum",
            "VALUE": "sum"
        },
        margins=True,
        margins_name="TOTAL"
    ).reset_index()

 
    total_cts = round(df_data['CTS'].sum(), 2)
    total_value = round(df_data['VALUE'].sum(), 2)
    avg_value_cts = round(total_value / total_cts, 2)
    

    # CTS %
    pivot5["CTS %"] = (
        pivot5["CTS"] / total_cts * 100
    ).round(2)

    # VALUE %
    pivot5["VALUE %"] = (
        pivot5["VALUE"] / total_value * 100
    ).round(2)

    # AVG VALUE/CTS
    pivot5["AVG VALUE/CTS"] = (
        pivot5["VALUE"] / pivot5["CTS"]
    ).round(2)
    
     # Round numeric columns to 2 decimals
    pivot5[["CTS", "CTS %", "VALUE", "VALUE %", "AVG VALUE/CTS"]] = (
    pivot5[["CTS", "CTS %", "VALUE", "VALUE %", "AVG VALUE/CTS"]].round(2)
    )

    pivot5["DTC SHAPE-ID"] = pd.to_numeric(
    pivot5["DTC SHAPE-ID"],
    errors="coerce"
    )

    # SORTING
    sort_df5 = (
        df_round[[
            "DTC SHAPE",
            "DTC SHAPE-ID",
        ]]
        .drop_duplicates()
    )

    pivot5 = pivot5.sort_values(
        by=[
            "DTC SHAPE-ID"
        ],
        na_position="last"
    )

    # COLUMN ORDER
    pivot5 = pivot5[[
        "DTC SHAPE",
        "CTS",
        "CTS %",
        "VALUE",
        "VALUE %",
        "AVG VALUE/CTS"
    ]]




# =====================================================
# SUMMARY DATA
# =====================================================

   ## summary_df = pd.DataFrame({
   #     "Metric": [
   #        "Selected Shape",
   #         "Total PCS",
   #         "Total CTS",
   #         "Total VALUE",
   #         "AVG VALUE/CTS"
   #     ],
   #     "Value": [
   #         "ROUND",
   #         len(df_round),
   #         round(total_cts, 2),
   #         round(total_value, 2),
   #         round(avg_value_cts, 2)
   #     ]
   # })

    # =====================================================
    # DISPLAY
    # =====================================================

    with col1:

        st.markdown(
            "Round Mb Gd Status & Dtc Quality Report"
        )

        st.dataframe(
            pivot1,
            use_container_width=True,
            height=650
        )

    with col2:
        #st.subheader
        st.markdown(
            "Round Mb Size Gr Details"
        )

        st.dataframe(
            pivot2,
            use_container_width=True,
            height=650
        )

    with col3:

       # st.subheader(
       #     "Pivot Table 3 : DTC Shade"
       # )

        st.markdown(
            "Round Dtc Shade Details"
        )

        st.dataframe(
            pivot3,
            use_container_width=True,
            height=650
        )

    with col4:

        #st.subheader(
        #    "Pivot Table 4 : DTC Size"
        #)
        st.markdown(
            "Round Dtc Size Details"
        )

        st.dataframe(
            pivot4,
            use_container_width=True,
            height=650
        )
        
    with col5:

        #st.subheader(
        #    "Pivot Table 5 : DTC Shape"
        #)
        
        st.markdown(
            "All Dtc Shape Details"
        )
        
        st.dataframe(
            pivot5,
            use_container_width=True,
            height=650
        )
    
    
    #with col3:

    #    st.subheader("Summary")

        #st.dataframe(
        #    summary_df,
        #    use_container_width=True,
        #    height=300
        #)
  
    excel_file = export_to_excel(pivot1,pivot2,pivot3,pivot4,pivot5,selected_lot)
    file_name = f"{'_'.join(selected_lot)}.xlsx"

    st.download_button(
        label="Download Excel Report",
        data=excel_file,
        file_name=file_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

# ---------------- MAIN ---------------- #

def main():

    setup_page()

    #st.title("Sample Polish Working")

    mst = mst_upload()
    file = file_upload()

    if mst is not None and file is not None:

        try:

            # LOAD MASTER (cached)
            (
                df_dtc_quality,
                df_shape_mst,
                df_shade_mst,
                df_size_mst,
                df_mbgd_mst,
                df_mbgd_P08_mst,
                df_dtc_size_mst,
            ) = load_master(mst)

            # LOAD DATA (cached)
            df_data = load_data(file)

            # =========================
            # FILTER FIRST (IMPORTANT)
            # =========================
            selected_lot = lot(df_data)
            l1 = "This data given hereunder is an estimated subject to verification reconciliation and approval  by directors."
            # Dynamic Title
            if selected_lot:
                    st.markdown(f"{l1}\n\nMaterial Number - {', '.join(selected_lot)}")

            else:
                st.markdown("Sample Polish Working")

            if selected_lot:
                df_data = df_data[
                    df_data['LOT'].astype(str).isin(selected_lot)
                ]


            # =========================
            # MERGE AFTER FILTER
            # =========================

            df_data = dtc_shape(df_data, df_shape_mst)
            df_data = dtc_quality(df_data, df_dtc_quality)
            df_data = shade(df_data, df_shade_mst)
            df_data = mb_size(df_data, df_size_mst)
            df_data = mb_gdstatus(df_data, df_mbgd_mst, df_mbgd_P08_mst)
            df_data = dtc_size(df_data, df_dtc_size_mst=df_dtc_size_mst)
            #st.subheader("Filtered Data")
            #st.dataframe(df_data)
            
            #st.subheader("Filtered Data")
            #st.dataframe(df_data)

            # RUN PIVOT FUNCTION
            create_pivots(df_data,selected_lot)

            st.success("Data Loaded Successfully")

            
        except Exception as e:
            st.error(f"Error : {e}")

    else:
        st.warning("Please upload both files")
        


if __name__ == "__main__":
    main()
